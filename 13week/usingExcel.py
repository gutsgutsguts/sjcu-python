# pip install openpyxl

from openpyxl import Workbook, load_workbook
import random

wb = Workbook()
ws = wb.active
print(ws)

ws = wb.create_sheet("New-Sheet")
print(ws)

ws_list = wb.sheetnames
print(ws_list)

ws["A1"] = "kor"
ws["B1"] = "math"
ws["C1"] = "grade"

for i in range(2,12):
    kor = random.randint(10,100)
    math = random.randint(10, 100)
    ws.cell(row=i, column=1).value = kor
    ws.cell(row=i, column=2).value = math
    ws["C"+str(i)].value = "=IF(AVERAGE(A"+str(i)+":B"+str(i)+")>=70, \"A\", IF(AVERAGE(A"+str(i)+":B"+str(i)+")>=40, \"B\", \"C\"))"
    
wb.save(r"grade.xlsx")

wb = load_workbook("grade.xlsx")
ws_list = wb.sheetnames
print(ws_list)

ws = wb["New-Sheet"]
print(ws)

range = ws["A1:C11"]
print(range)

for cell in range:
    print(cell[0].value, cell[1].value, cell[2].value)
    

